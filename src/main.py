import re
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from kivy.metrics import dp
from kivy.uix.button import Button
from kivy.properties import ObjectProperty
from kivymd.app import MDApp
from kivy.uix.screenmanager import ScreenManager
from kivy.lang import Builder
from kivy.core.window import Window
from kivymd.toast import toast
from kivy.utils import get_color_from_hex
from kivy.uix.popup import Popup
from kivymd.uix.label import MDLabel
from kivymd.uix.textfield import MDTextField
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.button import MDFloatingActionButton
from kivymd.uix.button import MDFillRoundFlatButton
import numpy as np
from PIL import Image
from pylibdmtx import pylibdmtx
from pyzbar import pyzbar
from kivy.uix.camera import Camera
from kivy.clock import Clock
import firebase_admin
import openpyxl
from firebase_admin import credentials
from firebase_admin import db
from firebase_admin import auth
import win32com.client as win32
from firebase_admin import storage

Window.size = (310, 580)


# 71642225|21CB01795|       |OT1|50930638

class FileItem(MDBoxLayout):
    def __init__(self, filename, **kwargs):
        super(FileItem, self).__init__(**kwargs)
        self.orientation = "vertical"
        self.padding = (dp(10), dp(5), dp(10), dp(5))

        label = MDLabel(text=filename)
        self.add_widget(label)


class Application(MDApp):
    camera = ObjectProperty(None)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.ids = None
        self.user_email = ''
        self.path_to_save = ''
        self.user_psw = ''
        self.user_clinic = ''
        self.user_doc=''
        self.user_surgery=''
        self.excel_name = ''
        self.file_list = ObjectProperty(None)  # Προσθήκη του πεδίου file_list
        self.list = []
        self.verify_email = None
        self.number = ''
        self.description_1 = ''
        self.text_inputs = None
        self.inputs = None
        self.first_popup = None
        self.second_popup = None
        self.third_popup = None
        self.row = 8
        self.flag = 1
        self.notes = None
        self.sxolia = None
        self.wb = openpyxl.load_workbook('template.xlsx')
        self.ws = self.wb.active
        self.file_names = []

    # noinspection PyGlobalUndefined
    def build(self):
        global screen_manager
        screen_manager = ScreenManager()

        screen_manager.add_widget(Builder.load_file("welcomeScreen.kv"))
        screen_manager.add_widget(Builder.load_file("loginScreen.kv"))
        screen_manager.add_widget(Builder.load_file("mainScreen.kv"))
        screen_manager.add_widget(Builder.load_file("signUpScreen.kv"))
        return screen_manager

    #@staticmethod
    def sign_in(self, email, psw):
        validation = r'^[A-Za-z0-9._%+-]+@mediplat\.gr$'
        if email == '' or psw == '':
            toast("Error, you have to fill all the inputs")
        else:
            if re.match(validation, email):
                try:
                    user = auth.get_user_by_email(email)
                    user_id = user.uid
                    if psw == user_id:
                        screen_manager.switch_to(screen_manager.get_screen("mainScreen"))
                        self.user_email = email
                        self.user_psw = psw
                        toast("Πραγματοποιήθηκε είσοδος με επιτυχία!")
                    else:
                        toast("Λάθος στοιχεία λογαριασμού")
                except auth.UserNotFoundError:
                    toast("Δεν βρέθηκε χρήστης με το συγκεκριμένο email.")
            else:
                toast("Invalid email address")

    @staticmethod
    def sign_up(name, email, psw, conf_psw):
        validation = r'^[A-Za-z0-9._%+-]+@mediplat\.gr$'
        if name == '' or email == '' or psw == '' or conf_psw == '':
            toast("Error, you have to fill all the inputs")
        else:
            if re.match(validation, email):
                if psw == conf_psw:
                    try:
                        user = auth.create_user(
                            uid=psw,
                            email=email,
                            password=psw,
                        )
                        # Εάν ο κώδικας φθάσει εδώ, σημαίνει ότι ο κωδικός είναι έγκυρος
                        toast("Επιτυχής εγγραφή στην εφαρμογή MedOffice")
                        database = firebase_admin.db.reference('Users').child(name)
                        data = {'name': name, 'email': email, 'password': psw}
                        database.child(psw).set(data)
                    except ValueError as e:
                        # Πιάνουμε το exception ValueError που προκαλείται από τον κωδικό
                        toast("Ο κωδικός πρέπει να αποτελείται από τουλάχιστον 6 χαρακτήρες")
                else:
                    toast("The password that you have entered does not match")
            else:
                toast("Error you are not a MediPlat employee")

    @staticmethod
    def init_db():
        cred = credentials.Certificate('firebase-sdk.json')

        firebase_admin.initialize_app(cred, {
            'databaseURL': 'https://medoffice-6b1e8-default-rtdb.firebaseio.com'
        })
        database = firebase_admin.db.reference('REG')

    def show(self, qr):
        if qr == '':
            toast("Error, you have to enter a qr code")
        else:
            creds = [2]
            creds = self.split_code(qr)
            desc = self.search_desc(creds[0])
            toast("Found something in db")

    @staticmethod
    def split_code(num):
        qr_code = num
        ref, batch_lot_number = qr_code.split('|', 1)
        ref = ref.strip()
        batch_lot_number, dummy = batch_lot_number.split('|', 1)
        batch_lot_number = batch_lot_number.strip()
        dummy = dummy.strip()
        return ref, batch_lot_number, dummy

    @staticmethod
    def search_desc(num):
        ref = db.reference('REG')
        result = ref.child(num).get()

        description = None

        if result:
            description = result
        return description

    def show_popup(self):
        if not self.first_popup:
            content = MDBoxLayout(orientation='vertical', spacing=10, padding=10)
            colors = {
                "White": get_color_from_hex("#FFFFFF"),
                "Black": get_color_from_hex("#000000"),
                "Blue": get_color_from_hex("#0000FF"),
                "Green": get_color_from_hex("#00FF00"),  # Πράσινο χρώμα για τα hint
            }
            self.text_inputs = [
                MDTextField(hint_text="Νοσοκομείο", line_color_normal=colors["White"],
                            foreground_color=colors["White"]),
                MDTextField(hint_text="Ιατρός", line_color_normal=colors["White"], foreground_color=colors["White"]),
                MDTextField(hint_text="ΑΜ", line_color_normal=colors["White"], foreground_color=colors["White"]),
                MDTextField(hint_text="Ημερομηνία", line_color_normal=colors["White"],
                            foreground_color=colors["White"]),
                MDTextField(hint_text="Χειρουργείο", line_color_normal=colors["White"],
                            foreground_color=colors["White"]),
            ]
            self.user_doc = self.text_inputs[1]
            self.user_clinic = self.text_inputs[0]
            self.user_surgery = self.text_inputs[4]
            for text_input in self.text_inputs:
                content.add_widget(text_input)

            button_container = MDBoxLayout(orientation='horizontal', spacing=10, size_hint=(1, 0.2))
            start_scan_button = MDFillRoundFlatButton(text="Start Scan", size_hint=(0.5, 0.5), pos_hint={'right': 1})
            close_button = MDFillRoundFlatButton(text="Close", size_hint=(0.5, 0.5), pos_hint={'left': 1})
            start_scan_button.bind(on_release=self.show_second_popup)
            close_button.bind(on_release=self.close_popup)
            button_container.add_widget(close_button)
            button_container.add_widget(start_scan_button)
            content.add_widget(button_container)
            popup_width = Window.width * 1
            popup_height = Window.height * 1
            self.first_popup = Popup(
                title="Δημιουργία χρέωσης",
                content=content,
                size_hint=(None, None),
                size=(popup_width, popup_height),
                background_color=(99, 38, 44, 1),
                title_color="000000"
            )
        # Εμφάνιση του πρώτου popup
        self.first_popup.open()
        return self.text_inputs

    def close_popup(self, *args):
        if self.first_popup:
            # Εκκαθάριση των δεδομένων
            for text_input in self.text_inputs:
                text_input.text = ""
            self.first_popup.dismiss()

    def show_second_popup(self, *args):
        if self.inputs is None:
            pass
        elif self.inputs[0].text == '':
            pass
        elif self.inputs[0].text != '':
            desc = self.search_desc(self.inputs[0].text)
            self.excel_fill(self.ws, self.row, self.inputs[0].text, desc, self.inputs[1].text, 1)
            for text_input in self.inputs:
                text_input.text = ''
            self.row += 1
        colors = {
            "White": get_color_from_hex("#FFFFFF"),
            "Black": get_color_from_hex("#000000"),
            "Blue": get_color_from_hex("#0000FF"),
        }
        if self.first_popup:
            self.init_excel(self.text_inputs[0].text, self.text_inputs[1].text, self.text_inputs[2].text,
                            self.text_inputs[3].text, self.text_inputs[4].text)
            self.first_popup.dismiss()
        content = MDBoxLayout(orientation='vertical')
        popup_width = Window.width * 1
        popup_height = Window.height * 1
        self.second_popup = Popup(
            title="Ανάγνωση QR κωδικού",
            content=content,
            size_hint=(None, None),
            size=(popup_width, popup_height),
            background_color=(0, 0, 0, 0)
        )
        self.second_popup.open()
        self.camera = Camera(resolution=(320, 240), size_hint=(1, 10), play=True)
        button_box = MDBoxLayout(orientation='horizontal', size_hint=(1, None), height='48dp')
        next_button = MDFillRoundFlatButton(text="Επόμενο", size_hint=(0.5, 1))
        next_button.bind(on_release=self.next_scan)
        button_box.add_widget(next_button)
        add_button = MDFillRoundFlatButton(text="Προσθήκη", size_hint=(0.5, 1))
        add_button.bind(on_release=self.show_third_popup)
        button_box.add_widget(add_button)
        content.add_widget(button_box)
        content.add_widget(self.camera)
        Clock.schedule_interval(lambda dt: self.scan_qr_codes(self.camera), 0.5)
        self.notes = MDTextField(hint_text="Παρατηρήσεις υλικού", line_color_normal=colors["White"],
                                 foreground_color=colors["Blue"], pos_hint={"center_x": .5, "center_y": 4})
        content.add_widget(self.notes)
        self.sxolia = MDTextField(hint_text="Σχόλια", line_color_normal=colors["White"],
                                  foreground_color=colors["Blue"],
                                  pos_hint={"center_x": .5, "center_y": 5})
        content.add_widget(self.sxolia)

        finish_button = MDFillRoundFlatButton(text="Τέλος", size_hint=(0.5, 0.5), pos_hint={'right': 1})
        finish_button.bind(on_release=self.finish)
        content.add_widget(finish_button)
        left = MDFloatingActionButton(
            icon="icons/left_arrow.png",
            pos_hint={"x": 0.1, "y": 0.0}
        )
        content.add_widget(left)
        left.bind(on_release=self.return_to_first_popup)

    def show_third_popup(self, *args):
        if self.second_popup:
            self.second_popup.dismiss()
            self.camera.play = False
        content = MDBoxLayout(orientation='vertical')
        popup_width = Window.width * 1
        popup_height = Window.height * 1
        self.third_popup = Popup(
            title="Εισαγωγή κωδικού",
            content=content,
            size_hint=(None, None),
            size=(popup_width, popup_height),
            background_color=(0, 0, 0, 0)
        )
        self.third_popup.open()

        colors = {
            "White": get_color_from_hex("#FFFFFF"),
            "Black": get_color_from_hex("#000000"),
            "Blue": get_color_from_hex("#0000FF"),
        }
        self.inputs = [
            MDTextField(hint_text="REF", line_color_normal=colors["White"], foreground_color=colors["Blue"]),
            MDTextField(hint_text="LOT", line_color_normal=colors["White"], foreground_color=colors["Blue"]),
            MDTextField(hint_text="ΠΟΣΟΤΗΤΑ", line_color_normal=colors["White"], foreground_color=colors["Blue"]),
            MDTextField(hint_text="ΣΧΟΛΙΟ ΥΛΙΚΟΥ", line_color_normal=colors["White"], foreground_color=colors["Blue"]),
        ]
        button_box = MDBoxLayout(orientation='vertical', size_hint=(1, None), height='48dp')
        for input in self.inputs:
            content.add_widget(input)
        telos = MDFillRoundFlatButton(text="Τοποθέτηση", size_hint=(0.5, 1))
        telos.bind(on_release=self.topothetisi)
        button_box.add_widget(telos)
        content.add_widget(button_box)

    def scan_qr_codes(self, camera):
        if camera.play:
            if camera.texture is not None and camera.texture.pixels is not None:
                image_data = camera.texture.pixels
                width, height = camera.resolution
                image_data = np.frombuffer(image_data, np.uint8).reshape(height, width, 4)
                image = Image.fromarray(image_data)

                barcodes = pylibdmtx.decode(image)

                if barcodes:  # Check if codes are found
                    camera.play = False  # Stop the camera

                for barcode in barcodes:
                    code_content = barcode.data.decode("utf-8")
                    length = len(code_content)
                    if length < 39:
                        self.handle_code_128(code_content, self.row)
                    else:
                        self.handle_qr_code(code_content, self.row)
            else:
                pass
        else:
            pass

    '''
    def scan_qr_codes(self, camera):

        if camera.play:
            if camera.texture is not None and camera.texture.pixels is not None:
                image_data = camera.texture.pixels
                width, height = camera.resolution
                image_data = np.frombuffer(image_data, np.uint8).reshape(height, width, 4)
                image = Image.fromarray(image_data)
                barcodes = pyzbar.decode(image)

                if barcodes:  # Check if QR codes are found
                    camera.play = False  # Stop the camera

                for barcode in barcodes:
                    qr_content = barcode.data.decode("utf-8")
                    length = len(qr_content)
                    if length < 39:
                        self.handle_code_128(qr_content, self.row)
                    else:
                        self.handle_qr_code(qr_content, self.row)
            else:
                pass
        else:
            pass
    '''

    def handle_code_128(self, qr_content, row):
        print(qr_content)
        ref = qr_content[2:16]
        lot = qr_content[-8:]
        print("Ref:" + ref)
        print("Lot:" + lot)
        ref_node = db.reference('REG').child(ref).get()
        if ref_node is not None:
            self.number, self.description_1 = next(iter(ref_node.items()))
            self.excel_fill(self.ws, self.row, self.number, self.description_1, lot, 0)
        else:
            toast("Δεν βρέθηκε το υλικό στην βάση δεδομένων")

    def handle_qr_code(self, qr_content, row):
        self.camera.play = False
        creds = [2]
        creds = self.split_code(qr_content)
        desc = self.search_desc(creds[0])
        self.excel_fill(self.ws, row, creds[0], desc, creds[1], 0)

    def return_to_first_popup(self, *args):
        if self.second_popup:
            self.second_popup.dismiss()
            self.camera.play = False
        self.show_popup()

    def topothetisi(self, *args):
        if self.third_popup:
            self.camera.play = False
            self.third_popup.dismiss()
            self.show_second_popup()
            self.camera.play = True

    def excel_fill(self, ws, row, ref, description, lot, type):
        if type == 0:
            toast("Ref:" + ref + "Lot:" + lot)
            ws.cell(row=row, column=1).value = ref
            # Unmerge the cell range
            merged_range = None
            if ws.merged_cells.ranges:
                for merged_range in ws.merged_cells.ranges:
                    if ws.cell(row=row, column=2).coordinate in merged_range:
                        break
            if merged_range:
                ws.unmerge_cells(str(merged_range))

                # Set the values for the individual cells within the merged cell range
                ws.cell(row=row, column=2).value = description
                ws.cell(row=row, column=3).value = lot
                ws.cell(row=row, column=4).value = 1
                # ws.cell(row=row, column=5).value = self.notes.text

                # Merge the cell range again
                self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
        elif type == 1:
            toast("Ref:" + ref + "Lot:" + lot)
            ws.cell(row=row, column=1).value = ref
            # Unmerge the cell range
            merged_range = None
            if ws.merged_cells.ranges:
                for merged_range in ws.merged_cells.ranges:
                    if ws.cell(row=row, column=2).coordinate in merged_range:
                        break
            if merged_range:
                ws.unmerge_cells(str(merged_range))

                # Set the values for the individual cells within the merged cell range
                ws.cell(row=row, column=2).value = description
                ws.cell(row=row, column=3).value = lot
                ws.cell(row=row, column=4).value = self.inputs[2].text
                ws.cell(row=row, column=5).value = self.inputs[3].text

                # Merge the cell range again
                self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)

    def init_excel(self, clinic, doc, am, date, surgery):
        self.ws['B2'] = clinic
        self.ws['B3'] = doc
        self.ws['B4'] = surgery
        self.ws['B5'] = am
        self.ws['B6'] = date
        return self.ws, self.wb

    def finish(self, instance):
        self.camera.play = True
        self.ws.cell(row=31, column=2).value = self.sxolia.text
        self.second_popup.dismiss()
        clinic = self.ws['B2'].value
        am = self.ws['B5'].value
        surgery = self.ws['B4'].value
        filename = f"{am}_{surgery}.xlsx"
        file_path = os.path.join("CHARGES", filename)
        print("file path:"+file_path)
        self.path_to_save = file_path
        self.excel_name = filename
        self.wb.save(file_path)
        for text_input in self.text_inputs:
            text_input.text = ""
        self.row = 8
        i = 0
        for i in range(8, 45):
            self.ws.cell(row=self.row+1, column=1).value = ''
            self.ws.cell(row=self.row + 1, column=2).value = ''
            self.ws.cell(row=self.row + 1, column=3).value = ''
            self.ws.cell(row=self.row + 1, column=4).value = ''
        self.camera.play = False

        #self.save_excel_file(self.user_email, filename, file_path)

        # Προσθήκη του ονόματος αρχείου στη λίστα
        self.file_names.append(filename)

        # Αναζήτηση του MDGridLayout με το id file_list στο mainScreen.kv
        scroll_view = self.root.get_screen('mainScreen').ids.scroll_view
        file_list = scroll_view.children[0]

        # Δημιουργία κουμπιού για το αρχείο και προσθήκη του στο GridLayout
        button = Button(text=filename, size_hint=(1, None), height=dp(48))
        button.bind(
            on_release=lambda btn: self.show_options(filename))  # Σύνδεση της συνάρτησης file_clicked() με το κουμπί
        file_list.add_widget(button)

        file_list.bind(minimum_height=file_list.setter('height'))
        self.root.get_screen('mainScreen').ids.scroll_view.scroll_to(file_list.children[0])



    def next_scan(self, *args):
        self.camera.play = True
        self.ws.cell(row=self.row, column=5).value = self.notes.text
        self.notes.text = " "
        self.row += 1

    def show_options(self, filename):
        # Δημιουργία των μικρότερων labels
        label_larisa = Button(text="Λάρισα", on_release=lambda instance: self.Larisa_BackOffice(filename))
        label_athens = Button(text="Αθήνα", on_release=lambda instance: self.Athens_BackOffice(filename))

        # Δημιουργία του Popup με τα μικρότερα labels
        content = MDBoxLayout(orientation='vertical')
        content.add_widget(label_larisa)
        content.add_widget(label_athens)
        popup = Popup(content=content, title="Αποστολή < "+filename+" > προς:", size_hint=(None, None), size=(300, 200))
        popup.open()

    def Larisa_BackOffice(self, filename):

        directory = "CHARGER"
        recipient_emails = ['m.parara@mediplat.gr', 't.theodorou@mediplat.gr', 'p.stamelos@mediplat.gr', 'larisa@mediplat.gr']

        for root, dirs, files in os.walk(directory):
            for file in files:
                if file.endswith('.xlsx') and file.lower().startswith(filename.lower()):
                    attachment_path = os.path.join(root, file)
                    subject = 'Χρέωση από Νοσοκομείο:'+self.user_clinic
                    body = ''

                    self.send_email(subject, body, recipient_emails, attachment_path)
                    return  # Τερματίζουμε την αναζήτηση μετά την αποστολή του πρώτου αρχείου

        print(f'Δεν βρέθηκε αρχείο με το όνομα "{filename}".')

    def Athens_BackOffice(self, filename):
        attachment_folder = "CHARGES"
        recipient_emails = ['m.parara@mediplat.gr', 't.theodorou@mediplat.gr', 'v.tsaperos@mediplat.gr']

        attachment_path = os.path.abspath(attachment_folder)

        file_list = os.listdir(attachment_path)
        self.user_clinic=self.ws['B2'].value
        self.user_doc = self.ws['B3'].value
        self.user_surgery = self.ws['B4'].value
        for file in file_list:
            if file == filename:
                attachment_file = os.path.join(attachment_path, file)
                subject = 'Χρέωση:'+self.user_clinic+' ΑΜ:'+self.ws['B5'].value+" Ημ.:"+self.ws['B6'].value
                body = 'Το χειρουργείο είναι:' + self.user_surgery + ". Αφορά τον Ιατρό:" + self.user_doc+". Σχόλια:"+self.sxolia.text
                self.send_email(subject, body, recipient_emails, attachment_file)
                return

            print(f'Το αρχείο "{file}" δεν ταιριάζει με το όνομα "{filename}"')

        print(f'Δεν βρέθηκε αρχείο με το όνομα "{filename}".')
        self.user_clinic = ''
        self.user_doc = ''
        self.user_surgery = ''

    #Send email operation
    def send_email(self, subject, body, recipient_emails, attachment_path):
        try:
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.Subject = subject
            mail.Body = body
            mail.To = ";".join(recipient_emails)
            if attachment_path is not None:
                mail.Attachments.Add(attachment_path)
            mail.Send()
            print("Email εστάλη με επιτυχία.")
        except Exception as e:
            print(f"Σφάλμα κατά την αποστολή του email: {str(e)}")

    def search_function(self, search_text):
        scroll_view = self.root.get_screen('mainScreen').ids.scroll_view
        file_list = scroll_view.children[0]

        file_list.clear_widgets()

        if not search_text:
            for filename in self.file_names:
                button = Button(text=filename, size_hint=(1, None), height=dp(48))
                button.bind(on_release=lambda btn, name=filename: self.show_options(name))
                file_list.add_widget(button)
        else:
            matching_files = [filename for filename in self.file_names if search_text.lower() in filename.lower()]
            if matching_files:
                for filename in matching_files:
                    button = Button(text=filename, size_hint=(1, None), height=dp(48))
                    button.bind(on_release=lambda btn, name=filename: self.show_options(name))
                    file_list.add_widget(button)

            if file_list.children:
                scroll_view.scroll_to(file_list.children[0])
            else:
                file_list.add_widget(
                    MDLabel(text="Δεν βρέθηκαν αρχεία με αυτά τα κριτήρια", size_hint=(1, None), height=dp(48)))

        file_list.bind(minimum_height=file_list.setter('height'))
        if file_list.children:
            scroll_view.scroll_to(file_list.children[0])

    '''
    def search_function(self, search_text):
        scroll_view = self.root.get_screen('mainScreen').ids.scroll_view
        file_list = scroll_view.children[0]

        # Καθαρίστε το GridLayout πριν από την αναζήτηση
        file_list.clear_widgets()

        # Εάν το κείμενο αναζήτησης είναι κενό, εμφανίστε όλα τα labels
        if not search_text:
            for filename in self.file_names:
                button = Button(text=filename, size_hint=(1, None), height=dp(48))
                button.bind(on_release=lambda btn, name=filename: self.show_options(name))
                file_list.add_widget(button)
        else:
            # Αναζητήστε τα labels που ταιριάζουν με το κείμενο αναζήτησης
            matching_files = [filename for filename in self.file_names if search_text.lower() in filename.lower()]
            if matching_files:
                for filename in matching_files:
                    button = Button(text=filename, size_hint=(1, None), height=dp(48))
                    button.bind(on_release=lambda btn, name=filename: self.show_options(name))
                    file_list.add_widget(button)
                # Ελέγχουμε εάν υπάρχουν παιδιά στο file_list πριν προσπαθήσουμε να κάνουμε scroll
                if file_list.children:
                    scroll_view.scroll_to(file_list.children[0])
            else:
                # Αφαιρούμε όλα τα προηγούμενα παιδιά του file_list
                file_list.clear_widgets()
                # Προσθέτουμε ένα label για να εμφανίσουμε το μήνυμα "Δεν βρέθηκαν αρχεία"
                label = MDLabel(text="Δεν βρέθηκαν αρχεία με αυτά τα κριτήρια", size_hint=(1, None), height=dp(48))
                file_list.add_widget(label)

        file_list.bind(minimum_height=file_list.setter('height'))
        scroll_view.scroll_to(file_list.children[0])
        '''

    def save_excel_file(self, user_email, filename, file_path):
        # Καθορίστε τον φάκελο αποθήκευσης βάσει του email του χρήστη
        folder_path = f"users/{user_email}/"

        # Ανεβάστε το αρχείο Excel στο Firebase Storage
        bucket = storage.bucket()
        blob = bucket.blob(folder_path + filename)
        blob.upload_from_filename(file_path)


if __name__ == "__main__":
    app = Application()
    app.init_db()
    app.run()
